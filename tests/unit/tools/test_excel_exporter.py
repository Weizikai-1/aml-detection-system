"""
Excel 报告导出测试 (Task 7-1)

覆盖:
- 单报告导出（4个Sheet结构、字段完整性、风险等级着色）
- 批量导出
- 边界情况（空报告、空交易列表）
- 戒律M1: 数据完整性（不编造）
"""
import os
from typing import List

import pytest
from openpyxl import load_workbook

from tools.excel_exporter import ExcelExporter
from graph.state import STRReport, SuspiciousTransaction, Transaction


# ============================================================
# 夹具
# ============================================================
@pytest.fixture()
def export_dir(tmp_path):
    """临时导出目录"""
    d = tmp_path / "exports"
    d.mkdir()
    return str(d)


@pytest.fixture()
def sample_transaction() -> Transaction:
    """单笔交易"""
    return {
        "transaction_id": "T001",
        "from_account": "A001",
        "to_account": "A002",
        "amount": 45000.0,
        "timestamp": "2025-01-01 10:00:00",
        "transaction_type": "transfer",
        "remark": "测试备注",
    }


@pytest.fixture()
def sample_suspicious(sample_transaction) -> SuspiciousTransaction:
    """可疑交易"""
    return {
        "transaction": sample_transaction,
        "rule_hits": ["分拆转账", "大额交易"],
        "risk_score": 75,
        "evidence": ["分拆转账: 1小时内收到5笔转账", "大额交易: 45000元"],
        "graph_evidence": None,
        "llm_analysis": None,
        "llm_confidence": None,
        "is_false_positive": None,
        "community_id": None,
    }


@pytest.fixture()
def sample_report(sample_suspicious) -> STRReport:
    """完整STR报告样本"""
    return {
        "report_id": "STR-20250101-TEST0001",
        "report_date": "2025-01-01 12:00:00",
        "report_type": "初始报告",
        "primary_account": "A002",
        "related_accounts": ["A001"],
        "customer_profile": {
            "account_type": "个人",
            "risk_rating": "high",
            "monitoring_status": "active",
        },
        "suspicious_transactions": [sample_suspicious],
        "total_suspicious_amount": 45000.0,
        "suspicious_patterns": ["分拆转账(1笔)", "大额交易(1笔)"],
        "risk_level": "high",
        "analysis_summary": "账户A002存在可疑交易特征",
        "evidence_chain": ["分拆转账: 1小时内收到5笔转账", "大额交易: 45000元"],
        "disposal_suggestion": "列入重点监控名单",
        "compliance_status": "pending",
        "compliance_notes": None,
        "reviewer": None,
        "final_decision": None,
    }


@pytest.fixture()
def multi_reports(sample_suspicious) -> List[STRReport]:
    """多份报告（用于批量导出测试）"""
    reports = []
    for i in range(3):
        r = {
            "report_id": f"STR-20250101-TEST{i:04d}",
            "report_date": "2025-01-01 12:00:00",
            "report_type": "初始报告",
            "primary_account": f"A{i:03d}",
            "related_accounts": [],
            "customer_profile": {},
            "suspicious_transactions": [sample_suspicious],
            "total_suspicious_amount": 45000.0,
            "suspicious_patterns": ["分拆转账(1笔)"],
            "risk_level": "medium" if i == 0 else "high",
            "analysis_summary": f"账户A{i:03d}可疑",
            "evidence_chain": ["分拆转账: 1小时内收到5笔转账"],
            "disposal_suggestion": "持续观察",
            "compliance_status": "pending",
            "compliance_notes": None,
            "reviewer": None,
            "final_decision": None,
        }
        reports.append(r)
    return reports


# ============================================================
# 单报告导出测试
# ============================================================
@pytest.mark.unit
def test_export_single_report_creates_file(export_dir, sample_report):
    """导出单个报告应生成xlsx文件"""
    exporter = ExcelExporter(output_dir=export_dir)
    path = exporter.export_report(sample_report)
    assert os.path.exists(path)
    assert path.endswith(".xlsx")


@pytest.mark.unit
def test_export_single_report_has_four_sheets(export_dir, sample_report):
    """导出的Excel应包含4个Sheet"""
    exporter = ExcelExporter(output_dir=export_dir)
    path = exporter.export_report(sample_report)

    wb = load_workbook(path)
    sheet_names = wb.sheetnames
    assert "报告概要" in sheet_names
    assert "可疑交易明细" in sheet_names
    assert "证据链" in sheet_names
    assert "模式与处置" in sheet_names
    assert len(sheet_names) == 4


@pytest.mark.unit
def test_summary_sheet_contains_key_fields(export_dir, sample_report):
    """报告概要Sheet应包含关键字段"""
    exporter = ExcelExporter(output_dir=export_dir)
    path = exporter.export_report(sample_report)

    wb = load_workbook(path)
    ws = wb["报告概要"]
    # 把所有单元格内容拼起来检查
    all_text = " ".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value)
    assert "STR-20250101-TEST0001" in all_text
    assert "A002" in all_text
    assert "高" in all_text  # 风险等级中文
    assert "45,000.00" in all_text  # 金额（带千分位格式）


@pytest.mark.unit
def test_transactions_sheet_contains_all_txns(export_dir, sample_report):
    """可疑交易明细Sheet应包含所有可疑交易"""
    exporter = ExcelExporter(output_dir=export_dir)
    path = exporter.export_report(sample_report)

    wb = load_workbook(path)
    ws = wb["可疑交易明细"]
    # 表头在第3行，数据从第4行开始
    # 检查交易ID
    found_t001 = False
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[1] == "T001":  # 第2列是交易ID
            found_t001 = True
            assert row[2] == "A001"  # 付款账户
            assert row[3] == "A002"  # 收款账户
            assert row[4] == 45000.0  # 金额
            assert row[9] == 75  # 风险分
            break
    assert found_t001, "交易T001未在明细Sheet中找到"


@pytest.mark.unit
def test_evidence_sheet_contains_all_evidence(export_dir, sample_report):
    """证据链Sheet应包含所有证据"""
    exporter = ExcelExporter(output_dir=export_dir)
    path = exporter.export_report(sample_report)

    wb = load_workbook(path)
    ws = wb["证据链"]
    all_text = " ".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value)
    assert "分拆转账" in all_text
    assert "大额交易" in all_text


@pytest.mark.unit
def test_patterns_sheet_contains_summary(export_dir, sample_report):
    """模式与处置Sheet应包含摘要和处置建议"""
    exporter = ExcelExporter(output_dir=export_dir)
    path = exporter.export_report(sample_report)

    wb = load_workbook(path)
    ws = wb["模式与处置"]
    all_text = " ".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value)
    assert "账户A002存在可疑交易特征" in all_text
    assert "列入重点监控名单" in all_text


@pytest.mark.unit
def test_risk_level_coloring(export_dir, sample_report):
    """风险等级单元格应有对应颜色填充"""
    exporter = ExcelExporter(output_dir=export_dir)
    path = exporter.export_report(sample_report)

    wb = load_workbook(path)
    ws = wb["报告概要"]
    # 找到风险等级行
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "风险等级":
                # 同行下一列是值
                value_cell = ws.cell(row=cell.row, column=cell.column + 1)
                # 应有填充色（high=FFEB9C）
                assert value_cell.fill.start_color.rgb is not None
                return
    pytest.fail("未找到风险等级行")


@pytest.mark.unit
def test_risk_score_coloring_in_transactions(export_dir, sample_report):
    """交易明细中的风险分单元格应有颜色（75分=high=橙黄）"""
    exporter = ExcelExporter(output_dir=export_dir)
    path = exporter.export_report(sample_report)

    wb = load_workbook(path)
    ws = wb["可疑交易明细"]
    # 数据从第4行开始，风险分是第10列
    risk_cell = ws.cell(row=4, column=10)
    assert risk_cell.value == 75
    # 应有填充色
    assert risk_cell.fill.start_color.rgb is not None


# ============================================================
# 批量导出测试
# ============================================================
@pytest.mark.unit
def test_export_multiple_reports(export_dir, multi_reports):
    """批量导出应生成多个文件"""
    exporter = ExcelExporter(output_dir=export_dir)
    paths = exporter.export_reports(multi_reports)
    assert len(paths) == 3
    for p in paths:
        assert os.path.exists(p)
        assert p.endswith(".xlsx")


@pytest.mark.unit
def test_export_multiple_reports_distinct_names(export_dir, multi_reports):
    """批量导出的文件名应不同"""
    exporter = ExcelExporter(output_dir=export_dir)
    paths = exporter.export_reports(multi_reports)
    names = [os.path.basename(p) for p in paths]
    assert len(set(names)) == 3  # 全部唯一


# ============================================================
# 边界情况测试
# ============================================================
@pytest.mark.unit
def test_export_report_with_empty_transactions(export_dir):
    """空可疑交易列表应能正常导出"""
    report = {
        "report_id": "STR-EMPTY",
        "report_date": "2025-01-01 12:00:00",
        "report_type": "初始报告",
        "primary_account": "A999",
        "related_accounts": [],
        "customer_profile": {},
        "suspicious_transactions": [],
        "total_suspicious_amount": 0,
        "suspicious_patterns": [],
        "risk_level": "low",
        "analysis_summary": "无可疑交易",
        "evidence_chain": [],
        "disposal_suggestion": "无需处置",
        "compliance_status": "pending",
        "compliance_notes": None,
        "reviewer": None,
        "final_decision": None,
    }
    exporter = ExcelExporter(output_dir=export_dir)
    path = exporter.export_report(report)
    assert os.path.exists(path)

    wb = load_workbook(path)
    # 4个Sheet都应该存在
    assert len(wb.sheetnames) == 4


@pytest.mark.unit
def test_export_report_custom_path(export_dir, sample_report, tmp_path):
    """指定output_path时应保存到该路径"""
    custom_path = str(tmp_path / "custom.xlsx")
    exporter = ExcelExporter(output_dir=export_dir)
    result_path = exporter.export_report(sample_report, output_path=custom_path)
    assert result_path == custom_path
    assert os.path.exists(custom_path)


@pytest.mark.unit
def test_export_report_creates_output_dir(tmp_path, sample_report):
    """输出目录不存在时应自动创建"""
    new_dir = str(tmp_path / "new_dir" / "nested")
    exporter = ExcelExporter(output_dir=new_dir)
    path = exporter.export_report(sample_report)
    assert os.path.exists(path)
    assert os.path.exists(new_dir)


# ============================================================
# 数据完整性测试（戒律 M1）
# ============================================================
@pytest.mark.unit
def test_no_fabricated_data_in_export(export_dir, sample_report):
    """导出内容不应包含编造数据（所有字段来自报告）"""
    exporter = ExcelExporter(output_dir=export_dir)
    path = exporter.export_report(sample_report)

    wb = load_workbook(path)
    # 收集所有单元格文本
    all_text = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    all_text.append(str(cell))
    full_text = " ".join(all_text)

    # 报告中没有的字段不应出现
    assert "编造" not in full_text
    assert "假数据" not in full_text
    # 报告中的真实字段应该都在
    assert "STR-20250101-TEST0001" in full_text
    assert "A001" in full_text
    assert "A002" in full_text
