"""
STR 报告 Excel 导出工具

将可疑交易报告导出为多 Sheet 的 Excel 文件，便于分析师审阅。

戒律:
- M1: 所有数据来自真实报告字段，不编造
- M2: 每笔可疑交易附带规则命中和证据
- M3: 风险评分0-100
- M4: 证据链完整保留

Sheet 结构:
1. 报告概要 - 报告元信息、主涉案账户、风险等级
2. 可疑交易明细 - 每笔可疑交易及其规则命中、风险分、证据
3. 证据链 - 去重后的完整证据列表
4. 模式与处置 - 可疑模式、分析摘要、处置建议
"""
import os
from typing import List, Dict, Any, Optional
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from graph.state import STRReport


# ============================================================
# 样式定义
# ============================================================
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="305496")
SUBTITLE_FONT = Font(name="微软雅黑", size=11, bold=True)
NORMAL_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)

# 风险等级颜色（背景）
RISK_FILLS = {
    "critical": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # 红
    "high":     PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # 橙黄
    "medium":   PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # 浅黄
    "low":      PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # 绿
}
RISK_FONTS = {
    "critical": Font(name="微软雅黑", size=10, bold=True, color="9C0006"),
    "high":     Font(name="微软雅黑", size=10, bold=True, color="9C6500"),
    "medium":   Font(name="微软雅黑", size=10, color="7F6000"),
    "low":      Font(name="微软雅黑", size=10, color="006100"),
}

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _risk_cn(level: str) -> str:
    """风险等级转中文"""
    return {
        "critical": "极高",
        "high": "高",
        "medium": "中",
        "low": "低",
    }.get(level, level)


def _apply_header_row(ws, row: int, headers: List[str]):
    """应用表头样式"""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _set_column_widths(ws, widths: List[int]):
    """设置列宽"""
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_summary_sheet(wb: Workbook, report: STRReport):
    """Sheet 1: 报告概要"""
    ws = wb.create_sheet("报告概要", 0)

    # 标题
    ws.merge_cells("A1:D1")
    ws["A1"] = "可疑交易报告 - 报告概要"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    # 基本信息（两列布局：字段名 | 值）
    rows = [
        ("报告编号", report.get("report_id", "")),
        ("报告日期", report.get("report_date", "")),
        ("报告类型", report.get("report_type", "")),
        ("风险等级", _risk_cn(report.get("risk_level", ""))),
        ("主涉案账户", report.get("primary_account", "")),
        ("关联账户数", len(report.get("related_accounts", []))),
        ("关联账户列表", "、".join(report.get("related_accounts", []))),
        ("可疑交易数", len(report.get("suspicious_transactions", []))),
        ("可疑交易总金额", f"{report.get('total_suspicious_amount', 0):,.2f} 元"),
        ("可疑模式", "、".join(report.get("suspicious_patterns", []))),
    ]

    # 客户画像
    profile = report.get("customer_profile", {})
    if profile:
        rows.append(("--- 客户画像 ---", ""))
        rows.append(("账户类型", profile.get("account_type", "")))
        rows.append(("风险评级", _risk_cn(profile.get("risk_rating", ""))))
        rows.append(("监控状态", profile.get("monitoring_status", "")))

    # 合规状态
    rows.append(("--- 合规状态 ---", ""))
    rows.append(("合规审核", report.get("compliance_status", "pending")))
    if report.get("compliance_notes"):
        rows.append(("合规备注", report["compliance_notes"]))
    if report.get("reviewer"):
        rows.append(("审核人", report["reviewer"]))
    if report.get("final_decision"):
        rows.append(("最终结论", report["final_decision"]))

    start_row = 3
    for i, (label, value) in enumerate(rows):
        r = start_row + i
        # 字段名
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = BOLD_FONT
        c1.alignment = WRAP_ALIGN
        c1.border = THIN_BORDER
        # 值
        c2 = ws.cell(row=r, column=2, value=value)
        c2.font = NORMAL_FONT
        c2.alignment = WRAP_ALIGN
        c2.border = THIN_BORDER

        # 风险等级特殊着色
        if label == "风险等级":
            risk = report.get("risk_level", "")
            if risk in RISK_FILLS:
                c2.fill = RISK_FILLS[risk]
                c2.font = RISK_FONTS[risk]

    _set_column_widths(ws, [20, 60])


def _write_transactions_sheet(wb: Workbook, report: STRReport):
    """Sheet 2: 可疑交易明细"""
    ws = wb.create_sheet("可疑交易明细")

    # 标题
    ws.merge_cells("A1:J1")
    ws["A1"] = "可疑交易明细"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    headers = [
        "序号", "交易ID", "付款账户", "收款账户", "金额(元)",
        "交易时间", "交易类型", "备注", "命中规则", "风险分",
    ]
    _apply_header_row(ws, 3, headers)

    txns = report.get("suspicious_transactions", [])
    for i, s in enumerate(txns, start=1):
        row = 3 + i
        t = s.get("transaction", {})
        rule_hits = "、".join(s.get("rule_hits", []))
        risk_score = s.get("risk_score", 0)

        values = [
            i,
            t.get("transaction_id", ""),
            t.get("from_account", ""),
            t.get("to_account", ""),
            float(t.get("amount", 0)),
            t.get("timestamp", ""),
            t.get("transaction_type", ""),
            t.get("remark", ""),
            rule_hits,
            risk_score,
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=v)
            cell.font = NORMAL_FONT
            cell.alignment = WRAP_ALIGN
            cell.border = THIN_BORDER

        # 金额格式
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        # 风险分颜色
        risk_cell = ws.cell(row=row, column=10)
        if risk_score >= 85:
            risk_cell.fill = RISK_FILLS["critical"]
            risk_cell.font = RISK_FONTS["critical"]
        elif risk_score >= 70:
            risk_cell.fill = RISK_FILLS["high"]
            risk_cell.font = RISK_FONTS["high"]
        elif risk_score >= 50:
            risk_cell.fill = RISK_FILLS["medium"]
            risk_cell.font = RISK_FONTS["medium"]
        else:
            risk_cell.fill = RISK_FILLS["low"]
            risk_cell.font = RISK_FONTS["low"]

    _set_column_widths(ws, [6, 18, 18, 18, 14, 20, 12, 30, 20, 10])
    ws.freeze_panes = "A4"


def _write_evidence_sheet(wb: Workbook, report: STRReport):
    """Sheet 3: 证据链"""
    ws = wb.create_sheet("证据链")

    ws.merge_cells("A1:C1")
    ws["A1"] = "证据链（完整可追溯）"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["序号", "证据内容", "关联交易ID"]
    _apply_header_row(ws, 3, headers)

    # 汇总证据（来自报告级 evidence_chain）
    evidence_chain = report.get("evidence_chain", [])
    txns = report.get("suspicious_transactions", [])

    # 构造证据-交易映射
    evidence_txn_map: Dict[str, List[str]] = {}
    for s in txns:
        t_id = s.get("transaction", {}).get("transaction_id", "")
        for ev in s.get("evidence", []):
            evidence_txn_map.setdefault(ev, []).append(t_id)

    for i, ev in enumerate(evidence_chain, start=1):
        row = 3 + i
        related = evidence_txn_map.get(ev, [])
        related_str = "、".join(related[:5])  # 最多显示5个关联交易
        if len(related) > 5:
            related_str += f" 等{len(related)}笔"

        for col_idx, v in enumerate([i, ev, related_str], start=1):
            cell = ws.cell(row=row, column=col_idx, value=v)
            cell.font = NORMAL_FONT
            cell.alignment = WRAP_ALIGN
            cell.border = THIN_BORDER

    _set_column_widths(ws, [6, 80, 30])
    ws.freeze_panes = "A4"


def _write_patterns_sheet(wb: Workbook, report: STRReport):
    """Sheet 4: 模式与处置"""
    ws = wb.create_sheet("模式与处置")

    ws.merge_cells("A1:B1")
    ws["A1"] = "可疑模式分析与处置建议"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    rows = [
        ("可疑模式", "、".join(report.get("suspicious_patterns", []))),
        ("分析摘要", report.get("analysis_summary", "")),
        ("处置建议", report.get("disposal_suggestion", "")),
    ]

    start_row = 3
    for i, (label, value) in enumerate(rows):
        r = start_row + i
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = BOLD_FONT
        c1.alignment = Alignment(vertical="top", wrap_text=True)
        c1.border = THIN_BORDER
        c2 = ws.cell(row=r, column=2, value=value)
        c2.font = NORMAL_FONT
        c2.alignment = WRAP_ALIGN
        c2.border = THIN_BORDER
        # 行高自适应（手动设置较大行高）
        ws.row_dimensions[r].height = 80

    _set_column_widths(ws, [16, 80])


class ExcelExporter:
    """STR 报告 Excel 导出器"""

    def __init__(self, output_dir: str = None):
        """
        Args:
            output_dir: 默认输出目录，None时使用 EXPORTS_DIR
        """
        if output_dir is None:
            from config import EXPORTS_DIR
            output_dir = EXPORTS_DIR
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

    def export_report(
        self,
        report: STRReport,
        output_path: str = None,
    ) -> str:
        """
        导出单个报告为 Excel 文件

        Args:
            report: STR 报告
            output_path: 输出路径，None时自动生成

        Returns:
            实际保存路径
        """
        if output_path is None:
            report_id = report.get("report_id", "STR-UNKNOWN")
            output_path = os.path.join(self.output_dir, f"{report_id}.xlsx")

        wb = Workbook()
        # 移除默认 Sheet
        default_ws = wb.active
        wb.remove(default_ws)

        # 按顺序构建4个 Sheet
        _write_summary_sheet(wb, report)
        _write_transactions_sheet(wb, report)
        _write_evidence_sheet(wb, report)
        _write_patterns_sheet(wb, report)

        wb.save(output_path)
        return output_path

    def export_reports(
        self,
        reports: List[STRReport],
        output_dir: str = None,
    ) -> List[str]:
        """
        批量导出多个报告

        Args:
            reports: STR 报告列表
            output_dir: 输出目录，None时使用默认目录

        Returns:
            保存路径列表
        """
        out_dir = output_dir or self.output_dir
        os.makedirs(out_dir, exist_ok=True)

        # 戒律 P4: 单报告失败不影响其他报告（错误隔离）
        paths = []
        for report in reports:
            report_id = report.get("report_id", f"STR-{datetime.now().strftime('%Y%m%d%H%M%S')}")
            path = os.path.join(out_dir, f"{report_id}.xlsx")
            try:
                self.export_report(report, path)
                paths.append(path)
            except Exception as e:
                print(f"  [Excel批量导出] 报告 {report_id} 失败: {e}")
        return paths
